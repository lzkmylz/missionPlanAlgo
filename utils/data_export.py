"""
数据导出模块

功能：
- 支持导出为CSV
- 支持导出为JSON
- 支持导出为Excel（可选）
- 支持数据格式化
"""

import csv
import json
import os
from collections import OrderedDict
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Any, List, Optional, Union


class DataExportError(Exception):
    """数据导出错误"""
    pass


class DataExporter:
    """
    数据导出器

    支持将数据导出为多种格式
    """

    def __init__(self):
        """初始化导出器"""
        pass

    def _validate_data(self, data: Any) -> None:
        """
        验证数据有效性

        Args:
            data: 待验证的数据

        Raises:
            DataExportError: 数据无效时抛出
        """
        if data is None:
            raise DataExportError("数据不能为None")

        if not isinstance(data, list):
            raise DataExportError(f"数据必须是列表类型，当前类型: {type(data).__name__}")

        if len(data) == 0:
            raise DataExportError("数据不能为空列表")

    def _ensure_dir(self, path: str) -> None:
        """
        确保目录存在

        Args:
            path: 文件路径
        """
        Path(path).parent.mkdir(parents=True, exist_ok=True)

    def _format_value(self, value: Any) -> str:
        """
        格式化值为字符串

        Args:
            value: 任意值

        Returns:
            str: 格式化后的字符串
        """
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def _get_all_keys(self, data: List[Dict[str, Any]]) -> List[str]:
        """
        获取所有数据的键集合

        Args:
            data: 数据列表

        Returns:
            List[str]: 键列表
        """
        keys = set()
        for item in data:
            if isinstance(item, dict):
                keys.update(item.keys())
        return sorted(list(keys))

    def to_csv(self, data: List[Dict[str, Any]], path: str) -> None:
        """
        导出数据为CSV文件

        Args:
            data: 数据列表
            path: 输出文件路径

        Raises:
            DataExportError: 导出失败时抛出
        """
        self._validate_data(data)

        try:
            self._ensure_dir(path)

            # 获取所有字段名
            fieldnames = self._get_all_keys(data)

            if not fieldnames:
                raise DataExportError("数据中没有有效的字段")

            with open(path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=fieldnames,
                    extrasaction='ignore',
                    quoting=csv.QUOTE_MINIMAL
                )
                writer.writeheader()

                for row in data:
                    if isinstance(row, dict):
                        # 格式化每一行的值
                        formatted_row = {
                            k: self._format_value(v) for k, v in row.items()
                        }
                        writer.writerow(formatted_row)

        except DataExportError:
            raise
        except Exception as e:
            raise DataExportError(f"导出CSV失败: {e}")

    def to_json(
        self,
        data: List[Dict[str, Any]],
        path: str,
        indent: Optional[int] = None,
        ensure_ascii: bool = False
    ) -> None:
        """
        导出数据为JSON文件

        Args:
            data: 数据列表
            path: 输出文件路径
            indent: 缩进空格数（None表示不格式化）
            ensure_ascii: 是否转义非ASCII字符

        Raises:
            DataExportError: 导出失败时抛出
        """
        self._validate_data(data)

        try:
            self._ensure_dir(path)

            with open(path, 'w', encoding='utf-8') as f:
                json.dump(
                    data,
                    f,
                    indent=indent,
                    ensure_ascii=ensure_ascii,
                    default=self._json_serializer
                )

        except DataExportError:
            raise
        except Exception as e:
            raise DataExportError(f"导出JSON失败: {e}")

    def _json_serializer(self, obj: Any) -> Any:
        """
        JSON序列化器

        Args:
            obj: 待序列化的对象

        Returns:
            Any: 可序列化的值
        """
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if hasattr(obj, '__dict__'):
            return obj.__dict__
        raise TypeError(f"无法序列化类型: {type(obj).__name__}")

    def to_excel(
        self,
        data: List[Dict[str, Any]],
        path: str,
        sheet_name: str = "Sheet1"
    ) -> None:
        """
        导出数据为Excel文件

        Args:
            data: 数据列表
            path: 输出文件路径
            sheet_name: 工作表名称

        Raises:
            DataExportError: 导出失败时抛出
        """
        self._validate_data(data)

        try:
            # 尝试导入openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
            except ImportError:
                raise DataExportError(
                    "导出Excel需要openpyxl库，请安装: pip install openpyxl"
                )

            self._ensure_dir(path)

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # 获取所有字段名
            fieldnames = self._get_all_keys(data)

            if not fieldnames:
                raise DataExportError("数据中没有有效的字段")

            # 写入表头
            for col, field in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col, value=field)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                if isinstance(row_data, dict):
                    for col_idx, field in enumerate(fieldnames, 1):
                        value = row_data.get(field, "")
                        formatted_value = self._format_excel_value(value)
                        ws.cell(row=row_idx, column=col_idx, value=formatted_value)

            # 调整列宽
            for col_idx, field in enumerate(fieldnames, 1):
                max_length = len(field)
                for row_data in data:
                    if isinstance(row_data, dict):
                        value = str(row_data.get(field, ""))
                        max_length = max(max_length, len(value))
                # 设置列宽（最大50）
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width

            # 保存文件
            wb.save(path)

        except DataExportError:
            raise
        except Exception as e:
            raise DataExportError(f"导出Excel失败: {e}")

    def _format_excel_value(self, value: Any) -> Union[str, int, float]:
        """
        格式化Excel单元格值

        Args:
            value: 任意值

        Returns:
            Union[str, int, float]: 格式化后的值
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def export(
        self,
        data: List[Dict[str, Any]],
        path: str,
        format: Optional[str] = None,
        **kwargs
    ) -> None:
        """
        通用导出方法

        Args:
            data: 数据列表
            path: 输出文件路径
            format: 导出格式（自动检测）
            **kwargs: 额外参数

        Raises:
            DataExportError: 导出失败时抛出
        """
        # 自动检测格式
        if format is None:
            ext = Path(path).suffix.lower()
            format_map = {
                '.csv': 'csv',
                '.json': 'json',
                '.xlsx': 'excel',
                '.xls': 'excel'
            }
            format = format_map.get(ext)

            if format is None:
                raise DataExportError(f"无法自动检测文件格式: {ext}")

        # 根据格式调用对应方法
        if format == 'csv':
            self.to_csv(data, path)
        elif format == 'json':
            self.to_json(data, path, **kwargs)
        elif format in ('excel', 'xlsx'):
            self.to_excel(data, path, **kwargs)
        else:
            raise DataExportError(f"不支持的导出格式: {format}")
